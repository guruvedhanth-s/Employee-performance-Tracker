"""
Billing Service - New Version
Business logic for billing reports and calculations
Billing is done organization-wide grouped by product types with shortened team names
"""
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, date
from typing import Optional, List, Dict
from io import BytesIO
from app.models.billing import BillingReport, BillingDetail
from app.models.order import Order
from app.models.team import Team
from app.models.user import User
from app.schemas.billing import (
    BillingReportCreate,
    BillingReportResponse,
    BillingDetailResponse,
    BillingPreviewRequest,
    BillingPreviewResponse,
    BillingPreviewDetail
)
from fastapi import HTTPException

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def get_team_short_name(team_name: str) -> str:
    """
    Convert team name to shortened version for product types
    Washington -> WA, Florida -> FL, California -> CA, etc.
    """
    team_mapping = {
        'Washington': 'WA',
        'Florida': 'FL',
        'California': 'CA',
        'Utah': 'UT',
        'Michigan': 'MI',
        'Oregon': 'OR',
        'Texas': 'TX',
        'Georgia': 'GA',
        'Vietnam Team': 'VN',
        'GI Clearing': 'GI',
        'Regional Streamline': 'RS',
        'National Streamline': 'NS',
        'FIF': 'FIF'
    }
    return team_mapping.get(team_name, team_name[:2].upper())


def format_product_type(team_name: str, product_type: str) -> str:
    """
    Format product type with team prefix
    Example: Washington + Full Search -> "WA Direct Full Search"
    """
    short_name = get_team_short_name(team_name)
    # Remove any existing state prefix from product type
    product_clean = product_type.strip()
    return f"{short_name} {product_clean}"


def get_billing_reports(
    db: Session,
    org_id: int,
    billing_month: Optional[int] = None,
    billing_year: Optional[int] = None,
    status: Optional[str] = None
) -> List[BillingReportResponse]:
    """
    Get billing reports with filters (no team filtering - all reports are org-wide)
    """
    query = db.query(BillingReport).filter(
        BillingReport.org_id == org_id,
        BillingReport.team_id.is_(None)  # All new reports are org-wide
    )
    
    if billing_month:
        query = query.filter(BillingReport.billing_month == billing_month)
    if billing_year:
        query = query.filter(BillingReport.billing_year == billing_year)
    if status:
        query = query.filter(BillingReport.status == status)
    
    reports = query.order_by(
        BillingReport.billing_year.desc(),
        BillingReport.billing_month.desc()
    ).all()
    
    # Convert to response models with details
    result = []
    for report in reports:
        details = [
            BillingDetailResponse(
                id=d.id,
                state=d.state,
                productType=d.product_type,
                singleSeatCount=d.single_seat_count,
                onlyStep1Count=d.only_step1_count,
                onlyStep2Count=d.only_step2_count,
                totalCount=d.total_count
            )
            for d in report.details
        ]
        
        total_files = sum(d.total_count for d in report.details)
        
        response = BillingReportResponse(
            id=report.id,
            orgId=report.org_id,
            teamId=None,
            teamName="All Teams",
            billingMonth=report.billing_month,
            billingYear=report.billing_year,
            status=report.status,
            createdBy=report.created_by,
            createdByName=report.created_by_user.user_name if report.created_by_user else None,
            finalizedBy=report.finalized_by,
            finalizedByName=report.finalized_by_user.user_name if report.finalized_by_user else None,
            finalizedAt=report.finalized_at,
            createdAt=report.created_at,
            modifiedAt=report.modified_at,
            details=details,
            totalFiles=total_files
        )
        result.append(response)
    
    return result


def preview_billing_data(
    db: Session,
    org_id: int,
    request: BillingPreviewRequest
) -> BillingPreviewResponse:
    """
    Preview billing data before generating report
    Shows what will be included in the billing report
    Organization-wide, grouped by product type (team + product)
    """
    # Calculate date range for the billing month
    start_date = date(request.billing_year, request.billing_month, 1)
    if request.billing_month == 12:
        end_date = date(request.billing_year + 1, 1, 1)
    else:
        end_date = date(request.billing_year, request.billing_month + 1, 1)
    
    # Query ALL pending orders in the organization for this period
    orders = db.query(Order).filter(
        Order.org_id == org_id,
        Order.billing_status == 'pending',
        Order.entry_date >= start_date,
        Order.entry_date < end_date,
        Order.deleted_at.is_(None)
    ).all()
    
    if not orders:
        raise HTTPException(
            status_code=400,
            detail="No pending orders found for this period"
        )
    
    # Group by formatted product type (team + product)
    billing_data: Dict[str, Dict[str, int]] = {}
    
    for order in orders:
        # Get team name
        team = db.query(Team).filter(Team.id == order.team_id).first()
        if not team:
            continue
        
        # Format product type with team prefix
        formatted_product = format_product_type(team.name, order.product_type)
        
        if formatted_product not in billing_data:
            billing_data[formatted_product] = {
                'single_seat': 0,
                'only_step1': 0,
                'only_step2': 0
            }
        
        # Determine order type
        if order.step1_user_id and order.step2_user_id:
            billing_data[formatted_product]['single_seat'] += 1
        elif order.step1_user_id:
            billing_data[formatted_product]['only_step1'] += 1
        elif order.step2_user_id:
            billing_data[formatted_product]['only_step2'] += 1
    
    # Convert to preview details
    details = []
    for product_type, counts in sorted(billing_data.items()):
        total = counts['single_seat'] + counts['only_step1'] + counts['only_step2']
        details.append(
            BillingPreviewDetail(
                productType=product_type,
                singleSeatCount=counts['single_seat'],
                onlyStep1Count=counts['only_step1'],
                onlyStep2Count=counts['only_step2'],
                totalCount=total
            )
        )
    
    total_files = sum(d.total_count for d in details)
    teams_count = db.query(Team).filter(Team.org_id == org_id).count()
    
    return BillingPreviewResponse(
        billingMonth=request.billing_month,
        billingYear=request.billing_year,
        details=details,
        totalFiles=total_files,
        pendingOrdersCount=len(orders),
        teamsCount=teams_count
    )


def create_billing_report(
    db: Session,
    org_id: int,
    current_user_id: int,
    data: BillingReportCreate
) -> BillingReportResponse:
    """
    Create a new organization-wide billing report grouped by product types
    """
    # Check if report already exists for this period
    existing = db.query(BillingReport).filter(
        BillingReport.org_id == org_id,
        BillingReport.team_id.is_(None),  # Org-wide reports have null team_id
        BillingReport.billing_month == data.billing_month,
        BillingReport.billing_year == data.billing_year
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Billing report already exists for {data.billing_month}/{data.billing_year}"
        )
    
    # Calculate date range
    start_date = date(data.billing_year, data.billing_month, 1)
    if data.billing_month == 12:
        end_date = date(data.billing_year + 1, 1, 1)
    else:
        end_date = date(data.billing_year, data.billing_month + 1, 1)
    
    # Query pending orders for this period
    orders = db.query(Order).filter(
        Order.org_id == org_id,
        Order.billing_status == 'pending',
        Order.entry_date >= start_date,
        Order.entry_date < end_date,
        Order.deleted_at.is_(None)
    ).all()
    
    if not orders:
        raise HTTPException(
            status_code=400,
            detail="No pending orders found for this period"
        )
    
    # Create billing report (team_id is NULL for org-wide)
    report = BillingReport(
        org_id=org_id,
        team_id=None,  # Org-wide report
        billing_month=data.billing_month,
        billing_year=data.billing_year,
        status='draft',
        created_by=current_user_id
    )
    db.add(report)
    db.flush()  # Get report.id
    
    # Group orders by formatted product type
    billing_data: Dict[str, Dict[str, int]] = {}
    
    for order in orders:
        # Get team name
        team = db.query(Team).filter(Team.id == order.team_id).first()
        if not team:
            continue
        
        # Format product type with team prefix
        formatted_product = format_product_type(team.name, order.product_type)
        
        if formatted_product not in billing_data:
            billing_data[formatted_product] = {
                'single_seat': 0,
                'only_step1': 0,
                'only_step2': 0
            }
        
        # Determine order type
        if order.step1_user_id and order.step2_user_id:
            billing_data[formatted_product]['single_seat'] += 1
        elif order.step1_user_id:
            billing_data[formatted_product]['only_step1'] += 1
        elif order.step2_user_id:
            billing_data[formatted_product]['only_step2'] += 1
    
    # Create billing details
    for product_type, counts in billing_data.items():
        total = counts['single_seat'] + counts['only_step1'] + counts['only_step2']
        detail = BillingDetail(
            report_id=report.id,
            state="",  # Not used in new format
            product_type=product_type,
            single_seat_count=counts['single_seat'],
            only_step1_count=counts['only_step1'],
            only_step2_count=counts['only_step2'],
            total_count=total
        )
        db.add(detail)
    
    db.commit()
    db.refresh(report)
    
    # Return response
    return get_billing_report_by_id(db, report.id)


def get_billing_report_by_id(db: Session, report_id: int) -> BillingReportResponse:
    """Get a single billing report by ID"""
    report = db.query(BillingReport).filter(BillingReport.id == report_id).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Billing report not found")
    
    details = [
        BillingDetailResponse(
            id=d.id,
            state=d.state,
            productType=d.product_type,
            singleSeatCount=d.single_seat_count,
            onlyStep1Count=d.only_step1_count,
            onlyStep2Count=d.only_step2_count,
            totalCount=d.total_count
        )
        for d in report.details
    ]
    
    total_files = sum(d.total_count for d in report.details)
    
    return BillingReportResponse(
        id=report.id,
        orgId=report.org_id,
        teamId=None,
        teamName="All Teams",
        billingMonth=report.billing_month,
        billingYear=report.billing_year,
        status=report.status,
        createdBy=report.created_by,
        createdByName=report.created_by_user.user_name if report.created_by_user else None,
        finalizedBy=report.finalized_by,
        finalizedByName=report.finalized_by_user.user_name if report.finalized_by_user else None,
        finalizedAt=report.finalized_at,
        createdAt=report.created_at,
        modifiedAt=report.modified_at,
        details=details,
        totalFiles=total_files
    )


def finalize_billing_report(
    db: Session,
    report_id: int,
    current_user_id: int
) -> BillingReportResponse:
    """
    Finalize a billing report
    - Marks the report as 'finalized'
    - Updates all associated orders billing_status from 'pending' to 'done'
    """
    report = db.query(BillingReport).filter(BillingReport.id == report_id).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Billing report not found")
    
    if report.status == 'finalized':
        raise HTTPException(status_code=400, detail="Report is already finalized")
    
    # Calculate date range
    start_date = date(report.billing_year, report.billing_month, 1)
    if report.billing_month == 12:
        end_date = date(report.billing_year + 1, 1, 1)
    else:
        end_date = date(report.billing_year, report.billing_month + 1, 1)
    
    # Update all pending orders for this organization and period to 'done'
    updated_count = db.query(Order).filter(
        Order.org_id == report.org_id,
        Order.billing_status == 'pending',
        Order.entry_date >= start_date,
        Order.entry_date < end_date,
        Order.deleted_at.is_(None)
    ).update({
        Order.billing_status: 'done',
        Order.modified_at: datetime.now(),
        Order.modified_by: current_user_id
    }, synchronize_session=False)
    
    # Update report status
    report.status = 'finalized'
    report.finalized_by = current_user_id
    report.finalized_at = datetime.now()
    report.modified_at = datetime.now()
    
    db.commit()
    db.refresh(report)
    
    return get_billing_report_by_id(db, report.id)


def delete_billing_report(db: Session, report_id: int) -> None:
    """
    Delete a billing report (only if in draft status)
    """
    report = db.query(BillingReport).filter(BillingReport.id == report_id).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Billing report not found")
    
    if report.status == 'finalized':
        raise HTTPException(
            status_code=400,
            detail="Cannot delete finalized billing report"
        )
    
    db.delete(report)
    db.commit()


def export_billing_report_to_excel(db: Session, report_id: int) -> BytesIO:
    """
    Export billing report to Excel format - Product Type based
    """
    if not EXCEL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Please install openpyxl package."
        )
    
    # Get report
    report = db.query(BillingReport).filter(BillingReport.id == report_id).first()
    
    if not report:
        raise HTTPException(status_code=404, detail="Billing report not found")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Billing Report"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Header row for data table - Start at row 1 (removed report info section)
    header_row = 1
    headers = ['Team Name', 'Product Type', 'Total']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Group details by team (extract team from product_type)
    # Product types are formatted as "WA Full Search", "FL Update", etc.
    team_data = {}
    for detail in report.details:
        # Extract team code (first 2-3 chars before space)
        parts = detail.product_type.split(' ', 1)
        if len(parts) >= 2:
            team_code = parts[0]  # e.g., "WA", "FL", "CA"
            product = parts[1]     # e.g., "Full Search", "Update"
            
            if team_code not in team_data:
                team_data[team_code] = []
            team_data[team_code].append({
                'product': product,
                'count': detail.total_count
            })
    
    # Define team order based on common teams
    team_order = [
        'Florida', 'California', 'GI Clearing', 'Washington', 'Michigan', 
        'Colorado', 'Utah', 'Oregon', 'Regional Streamline', 'National Streamline', 
        'FIF', 'SCB & PD', 'Arizona', 'Texas', 'Pennsylvania', 'Ohio', 'Guam',
        'Georgia', 'Vietnam Team'
    ]
    
    # Map codes to full names
    team_name_map = {
        'FL': 'Florida',
        'CA': 'California',
        'GI': 'GI Clearing',
        'WA': 'Washington',
        'MI': 'Michigan',
        'CO': 'Colorado',
        'UT': 'Utah',
        'OR': 'Oregon',
        'RS': 'Regional Streamline',
        'NS': 'National Streamline',
        'FIF': 'FIF',
        'SC': 'SCB & PD',
        'SCB': 'SCB & PD',
        'AZ': 'Arizona',
        'TX': 'Texas',
        'PE': 'Pennsylvania',
        'PA': 'Pennsylvania',
        'OH': 'Ohio',
        'GU': 'Guam',
        'GA': 'Georgia',
        'VN': 'Vietnam Team'
    }
    
    total_files = 0
    current_row = header_row + 1
    
    # Sort teams by the order defined
    sorted_teams = []
    for team_code, products in team_data.items():
        team_full_name = team_name_map.get(team_code, team_code)
        sorted_teams.append((team_code, team_full_name, products))
    
    # Sort by team order
    def get_team_sort_key(item):
        team_full_name = item[1]
        try:
            return team_order.index(team_full_name)
        except ValueError:
            return 999  # Put unknown teams at end
    
    sorted_teams.sort(key=get_team_sort_key)
    
    # Write data rows grouped by team
    for team_code, team_full_name, products in sorted_teams:
        # Sort products alphabetically
        products.sort(key=lambda x: x['product'])
        
        # Write each product for this team
        for idx, product_data in enumerate(products):
            # Team name only on first row
            if idx == 0:
                ws.cell(row=current_row, column=1).value = team_full_name
            else:
                ws.cell(row=current_row, column=1).value = ""
            
            ws.cell(row=current_row, column=2).value = product_data['product']
            ws.cell(row=current_row, column=3).value = product_data['count']
            
            # Apply borders and alignment
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                if col == 3:  # Total column
                    cell.alignment = center_align
            
            total_files += product_data['count']
            current_row += 1
    
    # Total row
    ws.cell(row=current_row, column=1).value = "GRAND TOTAL"
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=2).value = total_files
    
    # Apply styling to total row
    for col in range(1, 3):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = border
        if col == 2:
            cell.alignment = center_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20  # Team Name
    ws.column_dimensions['B'].width = 30  # Product Type
    ws.column_dimensions['C'].width = 12  # Total
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
